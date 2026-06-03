#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
产科护理思政反思日记自动评改脚本
评分量规来源：《妇产科护理学》反思日记评价量规（2026年修订版）
支持 .docx / .txt 文件，批量处理指定文件夹，输出 Markdown 评改报告 + Excel 汇总表
"""

import os
import sys
import io
import re
import json
import datetime
import argparse

# Windows 控制台 UTF-8 输出
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─────────────────────────────────────
# 依赖检测 & 自动安装
# ─────────────────────────────────────
def ensure_package(pkg_import, pkg_install=None):
    import importlib, subprocess
    try:
        importlib.import_module(pkg_import)
    except ImportError:
        name = pkg_install or pkg_import
        print(f"[安装依赖] 正在安装 {name} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", name, "-q"])

ensure_package("docx", "python-docx")
ensure_package("openpyxl")
ensure_package("requests")

import docx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import requests as _requests

# ─────────────────────────────────────
# AI 创作检测（ai-detector skill / GPTHumanizer API）
# ─────────────────────────────────────
_AI_DETECTOR_API_URL = "https://detect.gpthumanizer.ai/api/detect_ai"
_AI_DETECTOR_TIMEOUT = 30

print("[AI创作检测] 使用 ai-detector (GPTHumanizer API)")

# ─────────────────────────────────────
# AI 创作识别（超过 40% 直接判定不合格）
# ─────────────────────────────────────
AI_TYPICAL_PHRASES = [
    "综上所述", "总而言之", "由此可见", "值得注意的是",
    "需要指出的是", "不言而喻", "诚然", "毋庸置疑",
    "从某种程度上来说", "在很大程度上", "具有十分重要的意义",
    "为...奠定了坚实基础", "是实现...的关键所在",
    "我们应当认识到", "在当今社会", "随着社会的不断发展",
    "护理人员应", "我们作为一名", "通过本次反思",
    "不仅...而且...", "既要...又要...",
    "首先...其次...最后...", "一方面...另一方面...",
]

AI_TRANSITION_WORDS = [
    "首先", "其次", "再次", "最后", "总之",
    "因此", "因而", "故而", "由此可见",
    "此外", "另外", "与此同时", "值得注意的是",
    "诚然", "不可否认", "尽管如此", "即便如此",
]

def detect_ai_content(text):
    """
    使用 ai-detector (GPTHumanizer API) 检测 AI 创作特征
    返回 (ai_probability 0-100, reasons:list)
    ai_probability > 40 → 直接判定不合格并警示

    若 API 调用失败，自动回退到内置规则检测。
    """
    # ── 优先使用 ai-detector API ──
    try:
        resp = _requests.post(
            _AI_DETECTOR_API_URL,
            json={"text": text[:5000]},
            timeout=_AI_DETECTOR_TIMEOUT
        )
        resp.raise_for_status()
        data = resp.json()

        ai_pct = min(100, max(0, int(data.get("ai_possibilities", 0) * 100)))
        predicted_class = data.get("class", "unknown")

        reasons = []
        probs = data.get("probabilities", {})

        if predicted_class == "ai":
            reasons.append("文本极可能直接由 AI 生成（分类=ai）")
        elif predicted_class == "ai_humanized":
            reasons.append("文本可能经过 AI 生成后人工润色（分类=ai_humanized）")
        elif predicted_class == "light_edited":
            reasons.append("文本可能为 AI 生成后少量修改（分类=light_edited）")

        ai_prob = probs.get("ai", 0)
        human_prob = probs.get("human", 0)
        humanized_prob = probs.get("ai_humanized", 0)

        if ai_prob > 0.5:
            reasons.append(f"AI 生成概率较高（{ai_prob:.0%}）")
        if humanized_prob > 0.3:
            reasons.append(f"AI 润色概率较高（{humanized_prob:.0%}）")
        if human_prob > 0.5:
            reasons.append("文本偏向人工撰写，AI 特征不明显")

        if ai_pct > 40 and not reasons:
            reasons.append(f"AI 可能性评分 {ai_pct}%，超过 40% 阈值")

        return ai_pct, reasons

    except Exception as _e:
        print(f"[AI创作检测] API 调用失败，回退到内置规则: {_e}")

    # ── 回退：内置规则检测（原逻辑，已针对中文护理反思日记优化）──
    sentences = re.split(r"[。！？；\n]", text)
    sentences = [s.strip() for s in sentences if len(s.strip()) >= 4]
    word_count = len(text.replace(" ", "").replace("\n", ""))
    reasons = []
    score = 0

    # 1. 句子长度均匀度（仅在文字较长时检查，短文跳过）
    if len(sentences) >= 6 and word_count >= 400:
        lengths = [len(s) for s in sentences]
        avg_len = sum(lengths) / len(lengths)
        variance = sum((l - avg_len) ** 2 for l in lengths) / len(lengths)
        cv = variance ** 0.5 / max(avg_len, 1)
        if cv < 0.28:
            score += 25
            reasons.append(f"句子长度异常均匀（变异系数={cv:.2f}），AI 特征明显")
        elif cv < 0.38:
            score += 12
            reasons.append(f"句子长度较均匀（变异系数={cv:.2f}）")

    # 2. AI 典型套话命中（仅在命中≥4个时计分）
    phrase_hits = [p for p in AI_TYPICAL_PHRASES if p in text]
    if len(phrase_hits) >= 5:
        score += 30
        reasons.append(f"AI 典型套话过多（命中 {len(phrase_hits)} 个：{', '.join(phrase_hits[:4])}...）")
    elif len(phrase_hits) >= 4:
        score += 18
        reasons.append(f"出现多个 AI 典型套话（命中 {len(phrase_hits)} 个）")

    # 3. 过渡词密集度（仅在真正密集时计分）
    trans_hits = [t for t in AI_TRANSITION_WORDS if t in text]
    unique_trans = len(set(trans_hits))
    if unique_trans >= 8:
        score += 15
        reasons.append(f"过渡词使用过于密集（{unique_trans} 种）")
    elif unique_trans >= 6:
        score += 8
        reasons.append(f"过渡词使用较密集（{unique_trans} 种）")

    # 4. 缺乏个人化表达
    first_person = ["我", "我的", "我觉得", "我认为", "我感到", "我体会到", "我意识到"]
    emotion_words = ["开心", "难过", "紧张", "害怕", "感动", "愧疚", "自豪", "遗憾",
                     "忐忑", "欣慰", "委屈", "纠结", "懊悔", "欣喜"]
    time_specific = ["那天", "上周", "昨天", "上午", "下午", "凌晨", "深夜"]
    fp_hits = sum(1 for w in first_person if w in text)
    em_hits = sum(1 for w in emotion_words if w in text)
    ts_hits = sum(1 for w in time_specific if w in text)
    personal_score = fp_hits + em_hits + ts_hits
    if personal_score == 0 and word_count >= 300:
        score += 15
        reasons.append("完全缺乏个人化表达（无第一人称/情绪词/具体时间描述）")
    elif personal_score <= 1 and word_count >= 400:
        score += 8
        reasons.append("个人化表达较少")

    # 5. 内容泛化度（专业词少）
    prof_hits = sum(1 for t in KEYWORDS["专业结合"] if t in text)
    if word_count >= 500 and prof_hits <= 3:
        score += 12
        reasons.append(f"专业内容偏少（{word_count}字仅命中 {prof_hits} 个专业词）")

    # 6. 段落结构过于整齐
    paragraphs = [p.strip() for p in text.split("\n") if len(p.strip()) >= 20]
    if len(paragraphs) >= 4:
        para_lengths = [len(p) for p in paragraphs]
        para_avg = sum(para_lengths) / len(para_lengths)
        para_std = (sum((l - para_avg) ** 2 for l in para_lengths) / len(para_lengths)) ** 0.5
        para_cv = para_std / max(para_avg, 1)
        if para_cv < 0.20:
            score += 12
            reasons.append(f"段落长度异常均匀（变异系数={para_cv:.2f}）")

    score = min(score, 100)
    return score, reasons


def check_ai_and_override(text, results, total, overall, comment):
    """
    调用 AI 检测，若 ai_pct > 40，强制覆写为不合格，并返回警示信息。
    返回 (total, overall, comment, ai_warning, ai_pct, ai_reasons)
    """
    ai_pct, ai_reasons = detect_ai_content(text)
    ai_warning = None
    if ai_pct > 40:
        ai_warning = (
            f"⚠️⚠️ AI 创作识别警示 ⚠️⚠️\n"
            f"本篇日记 AI 创作概率约为 {ai_pct}%（超过 40% 阈值），"
            f"存在以下 AI 生成特征：\n"
            + "\n".join(f"  - {r}" for r in ai_reasons) +
            f"\n根据评分规则，AI 创作内容视为未完成真实反思，直接判定为【不合格】。"
            f"请学生重新以真实个人经历撰写反思日记。"
        )
        total = 0
        overall = "不合格（AI创作）"
        comment = (
            f"经系统 AI 创作识别检测，本篇日记 AI 生成概率较高（{ai_pct}%），"
            "疑似直接使用 AI 工具生成而未进行真实个人反思。"
            "根据课程要求，反思日记必须基于真实临床经历撰写，"
            "请重新提交原创反思日记。\n"
            "AI 检测详情：\n" + '\n'.join(f"  - {r}" for r in ai_reasons)
        )
        for dim in results:
            results[dim]["score"] = 0
            results[dim]["level"] = "不合格"
            results[dim]["reason"] += " | AI创作检测未通过"
    return total, overall, comment, ai_warning, ai_pct, ai_reasons


# ─────────────────────────────────────
# 评分量规（来源：《妇产科护理学》反思日记评价量规）
# 满分100分，8个维度
# ─────────────────────────────────────
RUBRIC = {
    "情境呈现与问题意识": {
        "weight": 10,
        "desc": "能否从孕产妇、胎儿/新生儿、家庭和护理团队多视角识别问题",
        "criteria": {
            "优秀 (9-10)": "能清晰呈现临床情境，从多个视角（孕产妇/胎儿/家庭/团队）深入识别问题，视角全面",
            "良好 (7-8)": "能描述情境并识别主要问题，视角较丰富，能联系多角度分析",
            "合格 (5-6)": "基本呈现情境，有一定问题识别意识，视角较单一",
            "不合格 (0-4)": "情境描述模糊，缺乏问题识别，未体现多视角分析"
        }
    },
    "生命至上与母婴安全意识": {
        "weight": 15,
        "desc": "是否体现风险意识、急危重症识别、母婴安全优先",
        "criteria": {
            "优秀 (13-15)": "深刻体现母婴安全意识，能识别急危重症征象，风险意识强",
            "良好 (10-12)": "有较强的安全意识，能识别主要风险，安全优先观念明确",
            "合格 (7-9)": "有基本安全意识，对常见风险有一定认识",
            "不合格 (0-6)": "缺乏安全意识，未提及母婴保护，风险识别不足"
        }
    },
    "课程价值与职业认知": {
        "weight": 15,
        "desc": "是否理解妇产科护士职业使命、专业责任、团队协作和护理价值",
        "criteria": {
            "优秀 (13-15)": "深刻理解职业使命和专业价值，团队协作意识强，有职业认同感",
            "良好 (10-12)": "理解护士职业责任，团队协作意识较好，有较好职业认知",
            "合格 (7-9)": "对护理职业有基本认识，有团队协作意识",
            "不合格 (0-6)": "职业认知不足，缺乏团队意识或职业认同感"
        }
    },
    "人文关怀与共情能力": {
        "weight": 15,
        "desc": "是否理解孕产妇焦虑、疼痛、羞耻感、丧失体验等情绪",
        "criteria": {
            "优秀 (13-15)": "深刻理解孕产妇情绪，能共情焦虑/疼痛/羞耻等复杂情绪，有人文关怀行动",
            "良好 (10-12)": "有较好共情能力，能识别孕产妇情绪，提供关怀支持",
            "合格 (7-9)": "有基本共情意识，能认识到孕产妇情绪需求",
            "不合格 (0-6)": "缺乏共情意识，未关注孕产妇情绪感受"
        }
    },
    "专科护理思维与风险应对": {
        "weight": 15,
        "desc": "是否能结合妇产科护理知识分析问题并提出合理措施，体现专业判断、风险意识能力",
        "criteria": {
            "优秀 (13-15)": "能结合专业知识分析临床问题，提出合理护理措施，专业判断力强",
            "良好 (10-12)": "有较好的专业思维，能联系专业知识分析，有一定应对措施",
            "合格 (7-9)": "有基本专业意识，能结合课程知识分析",
            "不合格 (0-6)": "缺乏专业思维，专业与反思脱节"
        }
    },
    "科学精神与思辨思维": {
        "weight": 10,
        "desc": "是否能秉持严谨科学态度，思辨研判，循证反思优化护理举措",
        "criteria": {
            "优秀 (9-10)": "能循证反思，运用专业知识和科学思维优化护理措施，思辨能力强",
            "良好 (7-8)": "有一定科学态度，能反思并提出改进方向",
            "合格 (5-6)": "有简单反思，有基本科学意识",
            "不合格 (0-4)": "反思缺乏科学依据，流于表面"
        }
    },
    "职业认同与责任担当": {
        "weight": 10,
        "desc": "是否体现慎独精神、团队协作、守护生命的职业使命",
        "criteria": {
            "优秀 (9-10)": "深刻体现慎独精神和责任担当，团队协作好，守护生命使命感强",
            "良好 (7-8)": "有较好责任意识，慎独精神较好，团队意识强",
            "合格 (5-6)": "有基本责任意识，团队协作尚可",
            "不合格 (0-4)": "缺乏责任意识，未体现慎独精神"
        }
    },
    "自我反思与改进计划": {
        "weight": 10,
        "desc": "是否能反思自身不足，并提出具体可行的改进措施",
        "criteria": {
            "优秀 (9-10)": "深刻反思自身不足，提出具体可行的改进计划，有明确成长目标",
            "良好 (7-8)": "能识别不足，有改进方向和一定计划",
            "合格 (5-6)": "有基本反思，有简单改进意愿",
            "不合格 (0-4)": "缺乏反思，流于形式，无具体改进措施"
        }
    }
}

# 关键词库（按8维度分类）
KEYWORDS = {
    # 情境呈现
    "情境呈现与问题意识": [
        "孕产妇", "产妇", "家属", "家庭", "胎儿", "新生儿", "宝宝",
        "护理团队", "带教老师", "护士长", "助产士", "医生",
        "观察", "识别", "问题", "判断", "评估", "发现", "多视角",
    ],
    # 母婴安全
    "生命至上与母婴安全意识": [
        "生命", "安全", "风险", "危急", "急危重症", "胎儿窘迫", "窒息",
        "产后出血", "子痫", "前置胎盘", "胎盘早剥", "羊水栓塞",
        "宫缩乏力", "识别", "监测", "预防", "及时", "紧急", "抢救",
        "母婴安全", "生命至上", "安全意识", "风险识别", "优先",
    ],
    # 职业认知
    "课程价值与职业认知": [
        "职业使命", "专业责任", "团队协作", "护理价值", "职业素养",
        "慎独", "慎独精神", "责任担当", "使命", "责任", "奉献",
        "职业认同", "爱岗敬业", "守护生命", "团队合作", "协作",
        "职业精神", "专业价值", "护理伦理", "角色", "岗位职责",
    ],
    # 人文关怀
    "人文关怀与共情能力": [
        "人文关怀", "共情", "同理心", "理解", "焦虑", "紧张", "害怕",
        "疼痛", "宫缩痛", "会阴切口", "羞耻感", "丧失", "情绪",
        "心理", "安慰", "鼓励", "支持", "陪伴", "安抚", "解释",
        "倾听", "沟通", "温暖", "安心", "安全感", "关心", "呵护",
        "握着她的手", "握住", "共情能力", "情绪支持",
    ],
    # 专业结合
    "专业结合": [
        "产科", "孕产妇", "新生儿", "分娩", "产后", "哺乳", "宫缩", "胎儿",
        "助产", "产程", "高危", "妊娠", "母婴", "护理", "临床", "实习",
        "患者", "产妇", "家属", "剖宫产", "自然分娩",
        "肩难产", "会阴", "会阴侧切", "导尿", "胎头", "胎肩", "娩出",
        "胎儿窘迫", "宫口", "胎心", "胎膜", "羊水", "产程图", "缩宫素",
        "催产", "屈大腿法", "旋肩法", "压前肩法", "耻骨上加压", "娩肩",
        "第一产程", "第二产程", "第三产程", "宫缩乏力", "产后出血",
        "胎盘", "胎盘剥离", "软产道", "宫颈扩张", "先露", "头位", "臀位",
        "接生", "铺台", "器械", "无菌", "新生儿窒息", "Apgar", "阿普加",
        "脐带", "脐带绕颈", "母乳喂养", "早接触", "早吸吮", "皮肤接触",
        "导乐", "陪伴分娩", "分娩镇痛", "椎管内麻醉", "笑气",
        "阴道检查", "肛门检查", "内诊", "备皮", "灌肠",
        "产褥期", "恶露", "会阴护理", "乳房护理", "产后抑郁", "围产期",
        "孕周", "足月", "早产", "过期妊娠", "前置胎盘", "胎盘早剥",
        "妊娠期高血压", "妊娠期糖尿病", "子痫前期", "子痫",
    ],
    # 科学精神
    "科学精神与思辨思维": [
        "循证", "证据", "文献", "科学", "严谨", "批判", "反思",
        "分析", "研判", "思考", "优化", "改进", "总结", "复盘",
        "临床指南", "专家共识", "标准", "规范", "依据", "研究",
        "循证实践", "为什么", "分析", "思考", "评估", "判断",
    ],
    # 职业认同
    "职业认同与责任担当": [
        "慎独", "慎独精神", "责任", "担当", "使命", "守护",
        "团队协作", "团队合作", "配合", "协作", "沟通",
        "职业使命", "守护生命", "慎独", "责任担当",
        "职业精神", "专业精神", "敬畏生命", "爱岗敬业",
        "自我约束", "主动", "积极", "承担",
    ],
    # 反思与改进
    "自我反思与改进计划": [
        "反思", "不足", "改进", "成长", "提升", "认识",
        "今后", "以后", "下次", "如果", "我希望", "我将",
        "我应该", "我需要", "我打算", "具体措施", "做法",
        "多练习", "多请教", "加强", "提高", "学习",
        "自我审视", "自我批评", "总结", "教训", "收获", "成长",
        "做得不好的", "需要改进", "不足之处", "主要原因",
    ]
}


# ─────────────────────────────────────
# 文件读取
# ─────────────────────────────────────
def read_file(filepath):
    """读取 docx 或 txt 文件内容"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".docx":
        try:
            doc = docx.Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[读取失败: {e}]"
    elif ext == ".txt":
        for enc in ["utf-8", "gbk", "utf-8-sig"]:
            try:
                with open(filepath, encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        return "[读取失败: 无法识别文件编码]"
    else:
        return None

def extract_student_name(filepath, content):
    """从文件名或内容中提取学生姓名，优先从内容中提取"""
    for line in content.split("\n")[:10]:
        m = re.search(r"姓\s*名\s*[：:]\s*([\u4e00-\u9fa5]{2,4})", line)
        if m:
            return m.group(1)
        m2 = re.search(r"^([\u4e00-\u9fa5]{2,4})[，。]", line)
        if m2 and len(line) < 20:
            return m2.group(1)

    filename = os.path.splitext(os.path.basename(filepath))[0]
    m = re.search(r"[\u4e00-\u9fa5]{2,4}\s*$", filename)
    if m:
        return m.group().strip()
    m = re.search(r"^[\u4e00-\u9fa5]{2,4}", filename)
    if m:
        name = m.group()
        if name not in ["产时模块", "反思日记", "肩难产", "观察者"]:
            return name
    m = re.search(r"\d{6,12}([\u4e00-\u9fa5]{2,4})", filename)
    if m:
        return m.group(1)

    return filename


# ─────────────────────────────────────
# 语义评分（8维度专项评分函数）
# ─────────────────────────────────────

def _sem_situation(text):
    """情境呈现与问题意识（10分）"""
    kw = KEYWORDS["情境呈现与问题意识"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    perspectives = {
        "孕产妇/产妇": any(w in text for w in ["孕产妇", "产妇", "孕妇"]),
        "胎儿/新生儿": any(w in text for w in ["胎儿", "新生儿", "宝宝", "孩子"]),
        "家庭": any(w in text for w in ["家属", "家庭", "家人", "丈夫", "老公"]),
        "护理团队": any(w in text for w in ["带教", "护士长", "老师", "医生", "助产士", "团队"]),
    }
    perspective_count = sum(perspectives.values())

    if perspective_count >= 3 and len(hits) >= 5:
        score = 10
    elif perspective_count >= 2 and len(hits) >= 3:
        score = 7
    elif len(hits) >= 2 or perspective_count >= 1:
        score = 5
    else:
        score = 3

    details = {"多视角": f"{perspective_count}/4", "关键词命中": len(hits)}
    return score, details


def _sem_safety(text):
    """生命至上与母婴安全意识（15分）"""
    kw = KEYWORDS["生命至上与母婴安全意识"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    emergency = ["危急", "急危重症", "胎儿窘迫", "窒息", "产后出血", "子痫",
                 "羊水栓塞", "宫缩乏力", "抢救", "紧急", "风险"]
    emergency_hits = [w for w in emergency if w in text]

    response = ["监测", "识别", "及时", "报告", "处理", "评估", "预防", "关注"]
    response_hits = [w for w in response if w in text]

    if emergency_hits and len(hits) >= 4:
        score = 15
    elif emergency_hits or len(hits) >= 3:
        score = 12
    elif len(hits) >= 2:
        score = 9
    else:
        score = 6

    details = {"安全关键词": len(hits), "急危识别": len(emergency_hits), "应对行为": len(response_hits)}
    return score, details


def _sem_profession(text):
    """课程价值与职业认知（15分）"""
    kw = KEYWORDS["课程价值与职业认知"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    core = ["职业使命", "专业责任", "护理价值", "慎独", "责任担当",
            "守护生命", "职业认同", "职业精神"]
    core_hits = [w for w in core if w in text]

    if core_hits and len(hits) >= 4:
        score = 15
    elif core_hits or len(hits) >= 3:
        score = 11
    elif len(hits) >= 2:
        score = 8
    else:
        score = 5

    details = {"职业关键词": len(hits), "核心使命词": len(core_hits)}
    return score, details


def _sem_care(text):
    """人文关怀与共情能力（15分）"""
    kw = KEYWORDS["人文关怀与共情能力"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    emotion = ["焦虑", "紧张", "害怕", "疼痛", "羞耻感", "丧失",
               "担心", "不安", "恐惧", "委屈", "难过"]
    emotion_hits = [w for w in emotion if w in text]

    action = ["安慰", "鼓励", "支持", "陪伴", "安抚", "解释",
              "倾听", "沟通", "握着", "握住", "共情"]
    action_hits = [w for w in action if w in text]

    if emotion_hits and action_hits and len(hits) >= 4:
        score = 15
    elif emotion_hits or action_hits:
        score = 12
    elif len(hits) >= 3:
        score = 9
    elif len(hits) >= 1:
        score = 7
    else:
        score = 5

    details = {"关怀关键词": len(hits), "情绪词": len(emotion_hits), "关怀行动": len(action_hits)}
    return score, details


def _sem_professional(text):
    """专科护理思维与风险应对（15分）"""
    prof_kw = KEYWORDS["专业结合"]
    prof_hits = [w for w in prof_kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    evidence = ["循证", "分析", "判断", "评估", "措施", "处理", "应对",
                "专业", "风险", "识别", "依据", "原因"]
    evidence_hits = [w for w in evidence if w in text]

    procedure = ["我做了", "我协助", "我帮助", "我给", "我为", "我参与",
                  "操作", "处理", "实施", "配合", "执行", "护理"]
    proc_hits = [w for w in procedure if w in text]

    if prof_hits and evidence_hits and len(prof_hits) >= 5:
        score = 15
    elif prof_hits and (evidence_hits or len(prof_hits) >= 3):
        score = 12
    elif prof_hits or len(prof_hits) >= 2:
        score = 9
    else:
        score = 6

    details = {"专业术语": len(prof_hits), "循证分析": len(evidence_hits), "操作行为": len(proc_hits)}
    return score, details


def _sem_scientific(text):
    """科学精神与思辨思维（10分）"""
    kw = KEYWORDS["科学精神与思辨思维"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    if len(hits) >= 5:
        score = 10
    elif len(hits) >= 3:
        score = 7
    elif len(hits) >= 2:
        score = 5
    else:
        score = 3

    details = {"思辨关键词": len(hits)}
    return score, details


def _sem_responsibility(text):
    """职业认同与责任担当（10分）"""
    kw = KEYWORDS["职业认同与责任担当"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    core = ["慎独", "责任担当", "守护生命", "职业使命"]
    core_hits = [w for w in core if w in text]

    if core_hits and len(hits) >= 3:
        score = 10
    elif core_hits or len(hits) >= 2:
        score = 7
    elif len(hits) >= 1:
        score = 5
    else:
        score = 3

    details = {"责任关键词": len(hits), "核心词": len(core_hits)}
    return score, details


def _sem_reflection(text):
    """自我反思与改进计划（10分）"""
    kw = KEYWORDS["自我反思与改进计划"]
    hits = [w for w in kw if w in text]
    word_count = len(text.replace(" ", "").replace("\n", ""))

    concrete = ["具体措施", "具体做法", "多练习", "多请教", "加强",
                "我打算", "我将", "我计划", "今后我会"]
    concrete_hits = [w for w in concrete if w in text]

    quality = ["反思", "不足", "不足之处", "做得不好", "自我批评"]
    quality_hits = [w for w in quality if w in text]

    if concrete_hits and quality_hits and len(hits) >= 4:
        score = 10
    elif concrete_hits or (quality_hits and len(hits) >= 3):
        score = 7
    elif quality_hits or len(hits) >= 2:
        score = 5
    else:
        score = 3

    details = {"反思关键词": len(hits), "反思质量": len(quality_hits), "具体措施": len(concrete_hits)}
    return score, details


# ─────────────────────────────────────
# 维度评分入口
# ─────────────────────────────────────
def score_dimension(text, dim_name):
    """
    对单个维度打分，返回 (分数, 等级, 理由)
    """
    weight = RUBRIC[dim_name]["weight"]

    score_funcs = {
        "情境呈现与问题意识": _sem_situation,
        "生命至上与母婴安全意识": _sem_safety,
        "课程价值与职业认知": _sem_profession,
        "人文关怀与共情能力": _sem_care,
        "专科护理思维与风险应对": _sem_professional,
        "科学精神与思辨思维": _sem_scientific,
        "职业认同与责任担当": _sem_responsibility,
        "自我反思与改进计划": _sem_reflection,
    }

    func = score_funcs.get(dim_name)
    if func:
        score, details = func(text)
        score = max(0, min(score, weight))
        level = _ratio_to_level(score / weight)
        reason = " | ".join(f"{k}:{v}" for k, v in details.items())
        return score, level, reason

    # 回退：关键词计数
    keywords = KEYWORDS.get(dim_name, [])
    hits = [kw for kw in keywords if kw in text]
    density = len(hits) / max(len(text) / 100, 1)
    score = int(weight * 0.6) if density >= 0.3 else int(weight * 0.4)
    score = max(0, min(score, weight))
    return score, _ratio_to_level(score / weight), f"关键词命中 {len(hits)} 个"


def _ratio_to_level(ratio):
    """分数比例转等级"""
    if ratio >= 0.90:
        return "优秀"
    elif ratio >= 0.75:
        return "良好"
    elif ratio >= 0.60:
        return "合格"
    else:
        return "不合格"


def grade_essay(text):
    """对整篇日记评分，返回各维度结果和总分"""
    results = {}
    total = 0
    for dim in RUBRIC:
        score, level, reason = score_dimension(text, dim)
        results[dim] = {
            "score": score,
            "max_score": RUBRIC[dim]["weight"],
            "level": level,
            "reason": reason,
            "criteria_hint": RUBRIC[dim]["criteria"].get(
                next((k for k in RUBRIC[dim]["criteria"] if level in k), list(RUBRIC[dim]["criteria"].keys())[1]), ""
            )
        }
        total += score

    # 总体等级（满分100，按比例）
    if total >= 90:
        overall = "优秀"
    elif total >= 75:
        overall = "良好"
    elif total >= 60:
        overall = "合格"
    else:
        overall = "不合格"

    return results, total, overall


# ─────────────────────────────────────
# 生成评语（按8维度）
# ─────────────────────────────────────
COMMENTS_TEMPLATE = {
    "优秀": [
        "该同学的反思日记展现了极高的妇产科护理专业素养，能从多视角分析临床情境，将人文关怀、生命至上与专业思维有机融合，职业认同感强，科学精神突出，文字流畅，值得表扬。",
        "日记思想性强，情感真挚，能深刻理解孕产妇需求，展现了良好的慎独精神和责任担当，专业与思政融合自然，建议在今后实践中继续保持并发扬这种优秀的反思意识。"
    ],
    "良好": [
        "该同学能从多角度分析临床情境，有较好的人文关怀意识和专业思维能力，职业认知较清晰，科学态度较好，建议进一步深化反思，结合更多具体案例展开论述。",
        "日记整体质量较好，能较好地将专业知识和人文关怀应用于临床反思中，有一定的思辨意识和改进计划，但部分内容可以更加具体，建议结合典型案例深入分析。"
    ],
    "合格": [
        "该同学完成了基本反思要求，有一定的问题识别意识和人文关怀意识，职业认知和安全意识较明确，建议深入思考多视角分析的重要性，增加对具体临床场景的深度反思。",
        "日记基本达到课程要求，能联系专业知识和人文关怀进行反思，有基本的改进方向，建议进一步丰富临床情境描写，增加专业深度和人文温度的融合。"
    ],
    "不合格": [
        "该同学的反思日记在情境呈现、专业分析或人文关怀方面有所欠缺，建议认真回顾一次真实的临床经历，从多视角识别问题，结合专业知识进行深入反思，充实内容后再重新提交。",
        "日记内容较为简单，建议结合实习中的具体产科护理经历，从专业思维和人文关怀角度进行反思，注意体现对母婴安全的重视和对孕产妇情绪的理解，充实内容后再重新提交。"
    ]
}

def generate_comment(results, total, overall, student_name):
    """根据评分结果生成个性化评语"""
    base_key = re.sub(r"（[^）]+）", "", overall).strip()
    if base_key not in COMMENTS_TEMPLATE:
        base_key = overall
    base_comment = random.choice(COMMENTS_TEMPLATE[base_key])

    weakest = min(results, key=lambda d: results[d]["score"] / results[d]["max_score"])
    weakest_ratio = results[weakest]["score"] / results[weakest]["max_score"]
    strongest = max(results, key=lambda d: results[d]["score"] / results[d]["max_score"])

    suggestions = []
    if weakest_ratio < 0.6:
        dim_suggestions = {
            "情境呈现与问题意识": "建议增加对临床情境的详细描写，从孕产妇、胎儿、家庭、团队等多个视角识别和分析问题",
            "生命至上与母婴安全意识": "建议加强对母婴安全意识的反思，提高对急危重症征象的识别能力，时刻牢记安全优先原则",
            "课程价值与职业认知": "建议深化对妇产科护士职业使命和护理价值的认识，增强团队协作意识和专业责任感",
            "人文关怀与共情能力": "建议更多关注孕产妇的情绪体验，理解焦虑、疼痛等感受，并落实到具体的关怀行动中",
            "专科护理思维与风险应对": "建议加强专业知识与临床反思的结合，运用循证思维分析问题并提出合理护理措施",
            "科学精神与思辨思维": "建议增强反思的科学性和思辨深度，运用循证依据优化护理实践",
            "职业认同与责任担当": "建议强化慎独精神和责任担当意识，深刻认识守护生命的职业使命",
            "自我反思与改进计划": "建议深入反思自身不足，提出更具体、可行的改进措施和成长目标"
        }
        suggestions.append(dim_suggestions.get(weakest, ""))

    suggestion_str = "；".join(s for s in suggestions if s)
    if suggestion_str:
        full_comment = f"{base_comment}重点改进方向：{suggestion_str}。"
    else:
        full_comment = base_comment

    return full_comment, weakest, strongest


# ─────────────────────────────────────
# 输出评改报告（Markdown）
# ─────────────────────────────────────
def generate_report(student_name, filepath, content, results, total, overall, comment,
                    weakest, strongest, ai_warning=None, ai_pct=0, ai_reasons=None):
    """生成单个学生的 Markdown 评改报告"""
    now = datetime.datetime.now().strftime("%Y年%m月%d日 %H:%M")
    word_count = len(content.replace(" ", "").replace("\n", ""))

    lines = [
        f"# 妇产科护理反思日记评改报告",
        f"",
        f"| 项目 | 内容 |",
        f"|------|------|",
        f"| 学生姓名 | {student_name} |",
        f"| 文件 | {os.path.basename(filepath)} |",
        f"| 字数 | 约 {word_count} 字 |",
        f"| 评改时间 | {now} |",
        f"| **总分** | **{total}/100** |",
        f"| **综合等级** | **{overall}** |",
        f"| AI 创作概率 | {ai_pct}% ({'⚠️ 超标' if ai_pct > 40 else '✅ 正常'}) |",
    ]

    # AI 创作识别警示区块
    if ai_warning:
        lines += [
            f"",
            f"> ⚠️⚠️⚠️ **AI 创作识别未通过（概率 {ai_pct}%）** ⚠️⚠️⚠️",
            f">",
            f"> 经系统多维度检测，本篇日记存在明显 AI 生成特征：",
        ]
        for r in (ai_reasons or []):
            lines.append(f"> - {r}")
        lines += [
            f">",
            f"> **根据评分规则，AI 创作内容视为未完成真实反思，直接判定为【不合格】。**",
            f"> 请学生基于真实临床经历重新撰写反思日记。",
            f"",
            f"---",
            f"",
        ]

    # AI 检测详情（始终显示）
    lines += [
        f"",
        f"## AI 创作检测详情",
        f"",
        f"- **AI 创作概率**：{ai_pct}% （阈值 40%，{'⚠️ 已超标' if ai_pct > 40 else '✅ 未超标'}）",
    ]
    if ai_reasons:
        lines.append(f"- **检测原因**：")
        for r in ai_reasons:
            lines.append(f"  - {r}")
    else:
        lines.append(f"- **检测原因**：无明显 AI 生成特征")

    lines += [
        f"",
        f"---",
        f"",
        f"## 各维度评分",
        f"",
        f"| 维度 | 得分 | 满分 | 等级 | 评分依据 |",
        f"|------|------|------|------|----------|",
    ]

    for dim, r in results.items():
        tag = " ⭐" if dim == strongest else (" ⚠️" if dim == weakest else "")
        lines.append(f"| {dim}{tag} | {r['score']} | {r['max_score']} | {r['level']} | {r['reason']} |")

    lines += [
        f"",
        f"> ⭐ = 本次最佳维度  ⚠️ = 需要重点改进的维度",
        f"",
        f"---",
        f"",
        f"## 综合评语",
        f"",
        f"> {comment}",
        f"",
        f"---",
        f"",
        f"## 详细评分标准参考（来源：《妇产科护理学》反思日记评价量规）",
        f"",
    ]

    for dim, r in results.items():
        lines.append(f"### {dim}（满分 {r['max_score']} 分）")
        lines.append(f"")
        lines.append(f"**评价重点：** {RUBRIC[dim]['desc']}")
        lines.append(f"")
        for grade, desc in RUBRIC[dim]["criteria"].items():
            marker = "→ **本次等级**" if r["level"] in grade else ""
            lines.append(f"- **{grade}**：{desc} {marker}")
        lines.append(f"")

    return "\n".join(lines)


# ─────────────────────────────────────
# 生成 Excel 汇总表（8维度）
# ─────────────────────────────────────
def generate_excel(all_results, output_path):
    """生成班级成绩汇总 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩汇总"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2E6DB4")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    grade_fill = {
        "优秀": PatternFill("solid", fgColor="C6EFCE"),
        "良好": PatternFill("solid", fgColor="FFEB9C"),
        "合格": PatternFill("solid", fgColor="FFCC99"),
        "不合格": PatternFill("solid", fgColor="FFC7CE"),
    }

    dims = list(RUBRIC.keys())
    headers = ["序号", "姓名", "文件名"] + [f"{d}({RUBRIC[d]['weight']})" for d in dims] + ["总分", "等级", "综合评语"]
    col_widths = [6, 10, 22] + [13] * len(dims) + [8, 8, 42]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 35

    for idx, row_data in enumerate(all_results, 1):
        row_num = idx + 1
        values = [idx, row_data["name"], row_data["filename"]]
        for dim in dims:
            values.append(row_data["scores"].get(dim, 0))
        values += [row_data["total"], row_data["overall"], row_data["comment"]]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center if col not in [3, len(values)] else Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            if col == len(values) - 1:
                cell.fill = grade_fill.get(val, PatternFill())
        ws.row_dimensions[row_num].height = 50

    # 统计行
    stat_row = len(all_results) + 2
    ws.cell(row=stat_row, column=1, value="统计").font = Font(bold=True)
    ws.cell(row=stat_row, column=2, value=f"共 {len(all_results)} 人")

    if all_results:
        totals = [r["total"] for r in all_results]
        ws.cell(row=stat_row, column=len(dims) + 3, value="平均分").font = Font(bold=True)
        ws.cell(row=stat_row, column=len(dims) + 4, value=round(sum(totals) / len(totals), 1))
        ws.cell(row=stat_row, column=len(dims) + 4).font = Font(bold=True, color="CC0000")

        from collections import Counter
        dist = Counter(r["overall"] for r in all_results)
        dist_str = " | ".join(f"{k}:{v}人" for k, v in dist.items())
        ws.cell(row=stat_row + 1, column=1, value="等级分布")
        ws.cell(row=stat_row + 1, column=2, value=dist_str)

    # 第二个 Sheet：维度分析
    ws2 = wb.create_sheet("维度分析")
    ws2.cell(1, 1, "维度").font = Font(bold=True)
    ws2.cell(1, 2, "满分").font = Font(bold=True)
    ws2.cell(1, 3, "平均分").font = Font(bold=True)
    ws2.cell(1, 4, "最高分").font = Font(bold=True)
    ws2.cell(1, 5, "最低分").font = Font(bold=True)
    ws2.cell(1, 6, "优秀率%").font = Font(bold=True)

    for r2, dim in enumerate(dims, 2):
        scores = [r["scores"][dim] for r in all_results]
        levels = [r["levels"][dim] for r in all_results]
        excellent_rate = round(levels.count("优秀") / len(levels) * 100, 1) if levels else 0
        ws2.cell(r2, 1, dim)
        ws2.cell(r2, 2, RUBRIC[dim]["weight"])
        ws2.cell(r2, 3, round(sum(scores) / len(scores), 1) if scores else 0)
        ws2.cell(r2, 4, max(scores) if scores else 0)
        ws2.cell(r2, 5, min(scores) if scores else 0)
        ws2.cell(r2, 6, excellent_rate)

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"[✓] Excel 汇总表已保存：{output_path}")


# ─────────────────────────────────────
# 主程序
# ─────────────────────────────────────
def process_folder(folder_path, output_folder=None):
    """处理整个文件夹"""
    if not os.path.isdir(folder_path):
        print(f"[错误] 文件夹不存在：{folder_path}")
        sys.exit(1)

    if output_folder is None:
        output_folder = os.path.join(folder_path, "评改结果")
    os.makedirs(output_folder, exist_ok=True)

    files = [
        f for f in os.listdir(folder_path)
        if os.path.splitext(f)[1].lower() in (".docx", ".txt")
        and not f.startswith("~")
    ]

    if not files:
        print(f"[提示] 文件夹中未找到 .docx 或 .txt 文件：{folder_path}")
        return

    print(f"\n{'='*55}")
    print(f"  《妇产科护理学》反思日记自动评改系统")
    print(f"  评分量规：反思日记评价量规（2026年修订版）")
    print(f"  评分维度：8维度（情境/安全/职业/关怀/专业/思辨/责任/反思）")
    print(f"  满分：100分")
    print(f"{'='*55}")
    print(f"  输入文件夹：{folder_path}")
    print(f"  输出文件夹：{output_folder}")
    print(f"  待处理文件：{len(files)} 份")
    print(f"{'='*55}\n")

    all_results = []
    failed = []

    for i, filename in enumerate(files, 1):
        filepath = os.path.join(folder_path, filename)
        print(f"[{i}/{len(files)}] 正在评改：{filename}")

        content = read_file(filepath)
        if content is None or content.startswith("[读取失败"):
            print(f"  ⚠️  跳过（{content}）")
            failed.append(filename)
            continue

        if len(content.strip()) < 50:
            print(f"  ⚠️  跳过（内容过短，可能是空文件）")
            failed.append(filename)
            continue

        student_name = extract_student_name(filepath, content)
        results, total, overall = grade_essay(content)
        comment, weakest, strongest = generate_comment(results, total, overall, student_name)

        # AI 创作识别
        total, overall, comment, ai_warning, ai_pct, ai_reasons = (
            check_ai_and_override(text=content, results=results, total=total,
                                 overall=overall, comment=comment)
        )
        if ai_warning:
            print(f"  ⚠️  AI 创作识别触发！AI 概率={ai_pct}%，直接判定不合格")
            for r in ai_reasons:
                print(f"     - {r}")

        # 保存 Markdown 报告
        report_md = generate_report(
            student_name, filepath, content, results, total, overall,
            comment, weakest, strongest,
            ai_warning=ai_warning, ai_pct=ai_pct, ai_reasons=ai_reasons
        )
        report_filename = f"{os.path.splitext(filename)[0]}_评改报告.md"
        report_path = os.path.join(output_folder, report_filename)
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_md)

        # 收集汇总数据
        all_results.append({
            "name": student_name,
            "filename": filename,
            "total": total,
            "overall": overall,
            "comment": comment,
            "scores": {dim: results[dim]["score"] for dim in results},
            "levels": {dim: results[dim]["level"] for dim in results},
        })

        emoji = {"优秀": "[优秀]", "良好": "[良好]", "合格": "[合格]", "不合格": "[不合格]"}.get(overall, "")
        print(f"  {emoji} 总分 {total}/100  [{overall}]  -> {report_filename}")

    if all_results:
        excel_path = os.path.join(output_folder, "班级成绩汇总.xlsx")
        generate_excel(all_results, excel_path)

    print(f"\n{'='*55}")
    print(f"  评改完成！")
    print(f"  成功：{len(all_results)} 份  |  跳过：{len(failed)} 份")
    if all_results:
        avg = sum(r["total"] for r in all_results) / len(all_results)
        print(f"  班级平均分：{avg:.1f} 分")
        from collections import Counter
        dist = Counter(r["overall"] for r in all_results)
        for level in ["优秀", "良好", "合格", "不合格"]:
            if level in dist:
                print(f"  {level}：{dist[level]} 人")
    if failed:
        print(f"\n  以下文件处理失败：")
        for f in failed:
            print(f"    - {f}")
    print(f"\n  结果已保存至：{output_folder}")
    print(f"{'='*55}\n")


# ─────────────────────────────────────
# 命令行入口
# ─────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="《妇产科护理学》反思日记自动评改工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
评分维度（8维度，总分100分）：
  1. 情境呈现与问题意识（10分）
  2. 生命至上与母婴安全意识（15分）
  3. 课程价值与职业认知（15分）
  4. 人文关怀与共情能力（15分）
  5. 专科护理思维与风险应对（15分）
  6. 科学精神与思辨思维（10分）
  7. 职业认同与责任担当（10分）
  8. 自我反思与改进计划（10分）

等级标准：90+优秀 / 75+良好 / 60+合格 / 60以下不合格
注：完全使用AI写作直接判为不合格。

示例：
  python grader.py "C:\\Users\\Admin\\Desktop\\反思日记"
  python grader.py ./日记文件夹 --output ./评改输出
        """
    )
    parser.add_argument("folder", help="包含反思日记文件的文件夹路径（支持 .docx 和 .txt）")
    parser.add_argument("--output", "-o", default=None, help="评改结果输出文件夹（默认：输入文件夹下的'评改结果'子文件夹）")

    args = parser.parse_args()
    process_folder(args.folder, args.output)
